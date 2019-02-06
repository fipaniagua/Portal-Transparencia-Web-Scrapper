from selenium import webdriver
from time import sleep, time
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook

def save_table(file_name, page_source, first = True):
    if first:
        new_wb = Workbook()
    else:
        new_wb = load_workbook(file_name)
    new_wb.guess_types = False
    new_wa = new_wb.active
    soup = BeautifulSoup(page_source, 'html.parser')


    ths = soup.find_all('th')
    #ths = browser.find_elements_by_xpath('//*[@id="A2248:form-visualizar:datosplantilla_head"]//th'
    header = [th.string for th in ths]
    amount_colm = len(header)
    if first:
        new_wa.append(header)

    #raw_data = browser.find_elements_by_xpath('//*[@id="A2248:form-visualizar:datosplantilla_data"]//td')
    raw_data = soup.find_all('td')
    lines = [cell.string.replace(".","") if cell.string != None else cell.string for cell in raw_data]
    #lines = [cell.string for cell in raw_data]
    for row_index in range(int(len(raw_data)/amount_colm)):
        start = row_index*amount_colm
        end = start+amount_colm
        line = lines[start:end]
        new_wa.append(line)

    new_wb.save(file_name)
    return

def save_page(link, file_name):
    browser.get(link)
    first = True
    while True:
        sleep(1)
        save_table(file_name, browser.page_source, first)
        next = browser.find_elements_by_xpath('/html/body/div[1]/div/div[2]/div/div/div/div/div/div/div/div/div[4]/form/div/div[2]/div/div[2]/div/div[1]/span[4]')
        if len(next) > 0 and not "disabled" in next[0].get_attribute("class"):
            first = False
            next_btn = browser.find_element_by_xpath('//*[@id="A2248:form-visualizar:datosplantilla_paginator_top"]/span[4]/span')
            next_btn.click()
        else:
            break

path_to_chromedriver = "./chromedriver"
browser = webdriver.Chrome(executable_path = path_to_chromedriver)
browser.implicitly_wait(4)

'''
t = time()
save_page("https://www.portaltransparencia.cl/PortalPdT/pdtta/-/ta/AH001/PR/PHON/23448678", "test_1.xlsx")
print("t:" , time() - t)
'''


wb =  load_workbook(filename='./datos_links.xlsx', read_only=True)
ws = wb.active
rows = list(ws.rows)[1:]
i = 1
succes = 0
st = time()
for row in rows[i-1:]:
    cells = list(row)
    file_name = "./data/{0}_{1}_{2}_{3}_{4}_({5}).xlsx".format(cells[0].value, cells[1].value, cells[4].value, cells[3].value, cells[2].value, i)
    #file_name = "./data_muni/{6}/{0}_{1}_{2}_{3}_{4}_({5}).xlsx".format(cells[0].value, cells[1].value, cells[4].value, cells[3].value, cells[2].value, i, selected_muni)
    print(file_name)
    print("i:", i)
    succes +=1
    try:
        save_page(cells[6].value, file_name)
    except Exception as e:
        succes -=1
    i += 1
    print("succes:", succes)
    print("time:", time()-st)
